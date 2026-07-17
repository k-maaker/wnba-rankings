#!/usr/bin/env python3
"""
WNBA Betting Market Rankings — recreation of inpredictable's methodology.

Methodology (per inpredictable, July 2016 post):
  * Reverse-engineer implied power rankings from point spreads + totals.
  * Weighted linear regression. Each game contributes rows weighted by recency:
        market lines:  weight = 1 / (elapsed_games + 0.25)
        game results:  weight = 1 / (elapsed_games + 3.50)
    where elapsed_games = number of games that team has played since that game.
  * Home court advantage estimated within the regression.
  * GPF  = points favored vs. league-average opponent, neutral court
    oGPF = offensive component (points scored vs. average)
    dGPF = defensive component (points allowed vs. average; positive = good D)
    GOU  = implied over/under vs. an average opponent = 2*mu + oGPF - dGPF

Data source: ESPN public scoreboard API (scores + spread + over/under).
State: data/games.csv (accumulated lines & results), data/history.csv (daily
snapshots used for sparklines and "last week" ranks).

Usage:
  python wnba_rankings.py                  # fetch new dates, model, render
  python wnba_rankings.py --backfill       # re-scan whole season for odds/results
  python wnba_rankings.py --no-fetch       # model + render from stored data only
  python wnba_rankings.py --test           # synthetic-data sanity check
"""

import argparse
import datetime as dt
import json
import os
import re
import sys
import urllib.request
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd

# ----------------------------------------------------------------------------
# Config
# ----------------------------------------------------------------------------
BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "data")
OUT_DIR = os.path.join(BASE, "docs")   # GitHub Pages serves /docs
GAMES_CSV = os.path.join(DATA_DIR, "games.csv")
HISTORY_CSV = os.path.join(DATA_DIR, "history.csv")

SEASON_START = os.environ.get("WNBA_SEASON_START", "2026-05-08")
EASTERN = ZoneInfo("America/New_York")
BDB_DIR = os.path.join(DATA_DIR, "bigdataball")

ODDS_DENOM = 0.25     # market-line weight: 1/(elapsed + 0.25)
RESULT_DENOM = 3.50   # game-result weight: 1/(elapsed + 3.5)
SPARK_N = 30          # points shown in sparklines

SCOREBOARD_URL = (
    "https://site.api.espn.com/apis/site/v2/sports/basketball/wnba/"
    "scoreboard?dates={date}"
)

# ESPN abbreviation -> inpredictable-style abbreviation
ABBR = {
    "NY": "NYL", "NYL": "NYL",
    "MIN": "MIN",
    "LV": "LVA", "LVA": "LVA",
    "GS": "GSV", "GSV": "GSV",
    "ATL": "ATL",
    "DAL": "DAL",
    "IND": "IND",
    "LA": "LAS", "LAS": "LAS",
    "PHX": "PHO", "PHO": "PHO",
    "TOR": "TOR",
    "CHI": "CHI",
    "WSH": "WAS", "WAS": "WAS",
    "SEA": "SEA",
    "POR": "PDX", "PDX": "PDX",
    "CONN": "CON", "CON": "CON",
}

TEAM_NAMES = {
    "NYL": "New York Liberty", "MIN": "Minnesota Lynx", "LVA": "Las Vegas Aces",
    "GSV": "Golden State Valkyries", "ATL": "Atlanta Dream", "DAL": "Dallas Wings",
    "IND": "Indiana Fever", "LAS": "Los Angeles Sparks", "PHO": "Phoenix Mercury",
    "TOR": "Toronto Tempo", "CHI": "Chicago Sky", "WAS": "Washington Mystics",
    "SEA": "Seattle Storm", "PDX": "Portland Fire", "CON": "Connecticut Sun",
}
FULLNAME_TO_ABBR = {v: k for k, v in TEAM_NAMES.items()}


def norm_team(abbr: str) -> str:
    return ABBR.get(abbr.upper(), abbr.upper())


# ----------------------------------------------------------------------------
# Fetching
# ----------------------------------------------------------------------------
def http_get_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def parse_spread_home(odds: dict, home: str, away: str):
    """Return the home-team spread (negative = home favored), or None."""
    # Preferred: parse 'details' like "NY -8.5" against team abbreviations.
    details = odds.get("details") or ""
    m = re.match(r"\s*([A-Z]+)\s+([+-]?\d+(\.\d+)?)\s*$", details)
    if m:
        fav, num = norm_team(m.group(1)), float(m.group(2))
        num = -abs(num)
        if fav == home:
            return num
        if fav == away:
            return -num
    if details.strip().upper() in ("EVEN", "PK", "PICK"):
        return 0.0
    # Fallback: ESPN numeric 'spread' is quoted from the home team's perspective.
    if odds.get("spread") is not None:
        try:
            return float(odds["spread"])
        except (TypeError, ValueError):
            pass
    return None


def parse_scoreboard(js: dict) -> list[dict]:
    rows = []
    for ev in js.get("events", []):
        comp = (ev.get("competitions") or [{}])[0]
        teams = comp.get("competitors") or []
        if len(teams) != 2:
            continue
        home = next((t for t in teams if t.get("homeAway") == "home"), None)
        away = next((t for t in teams if t.get("homeAway") == "away"), None)
        if not home or not away:
            continue
        h = norm_team(home["team"]["abbreviation"])
        a = norm_team(away["team"]["abbreviation"])
        if h not in TEAM_NAMES or a not in TEAM_NAMES:
            continue  # exhibition vs. national team, All-Star game, etc.
        status = (ev.get("status") or {}).get("type", {})
        final = bool(status.get("completed"))
        raw_date = ev.get("date", "")
        try:  # ESPN timestamps are UTC; convert to ET so late West Coast
            game_date = (dt.datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                         .astimezone(EASTERN).date().isoformat())
        except ValueError:
            game_date = raw_date[:10]
        row = {
            "game_id": ev.get("id"),
            "date": game_date,
            "home": h, "away": a,
            "home_score": int(home["score"]) if final and home.get("score") else np.nan,
            "away_score": int(away["score"]) if final and away.get("score") else np.nan,
            "spread_home": np.nan, "total": np.nan,
        }
        for odds in comp.get("odds") or []:
            sp = parse_spread_home(odds, h, a)
            if sp is not None:
                row["spread_home"] = sp
            if odds.get("overUnder") is not None:
                try:
                    row["total"] = float(odds["overUnder"])
                except (TypeError, ValueError):
                    pass
            if not np.isnan(row["spread_home"]) and not np.isnan(row["total"]):
                break
        rows.append(row)
    return rows


def load_games() -> pd.DataFrame:
    cols = ["game_id", "date", "home", "away",
            "home_score", "away_score", "spread_home", "total"]
    if os.path.exists(GAMES_CSV):
        df = pd.read_csv(GAMES_CSV, dtype={"game_id": str})
        df = df[df.home.isin(TEAM_NAMES) & df.away.isin(TEAM_NAMES)]
        return df[cols]
    return pd.DataFrame(columns=cols)


def merge_games(existing: pd.DataFrame, new_rows: list[dict]) -> pd.DataFrame:
    """Merge, never clobbering a stored line/score with a missing value."""
    df = existing.set_index("game_id") if len(existing) else pd.DataFrame(
        columns=[c for c in existing.columns if c != "game_id"]
    ).set_index(pd.Index([], name="game_id"))
    for r in new_rows:
        gid = str(r["game_id"])
        if gid in df.index:
            for k in ("home_score", "away_score", "spread_home", "total"):
                v = r.get(k, np.nan)
                if v is not None and not pd.isna(v):
                    df.loc[gid, k] = v
            df.loc[gid, ["date", "home", "away"]] = [r["date"], r["home"], r["away"]]
        else:
            df.loc[gid] = {k: r.get(k, np.nan) for k in df.columns}
    df = df.reset_index().rename(columns={"index": "game_id"})
    return df.sort_values("date").reset_index(drop=True)


def fetch_range(start: dt.date, end: dt.date) -> list[dict]:
    rows = []
    d = start
    while d <= end:
        url = SCOREBOARD_URL.format(date=d.strftime("%Y%m%d"))
        try:
            rows.extend(parse_scoreboard(http_get_json(url)))
        except Exception as e:  # noqa: BLE001 — log and keep going
            print(f"  ! {d}: {e}", file=sys.stderr)
        d += dt.timedelta(days=1)
    return rows


# ----------------------------------------------------------------------------
# BigDataBall xlsx ingestion (real closing lines, one file per season)
# ----------------------------------------------------------------------------
def parse_bigdataball_xlsx(path: str) -> pd.DataFrame:
    """Read a BigDataBall 'WNBA team feed' workbook into per-game rows with
    real opening/closing spreads and totals. Any file dropped in
    data/bigdataball/ is picked up automatically each run."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if re.match(r"^WNBA-\d{4}-TEAM$", s)), None
    ) or next(
        (s for s in wb.sheetnames
         if s.upper() not in ("METADATA", "TEAMS", "CONVERT DATE FORMAT")),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(h).replace("\n", " ").strip() if h else h for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(name):
        return idx.get(name)

    i_gid, i_date, i_team, i_venue, i_f = (
        col("GAME-ID"), col("DATE"), col("TEAMS"), col("VENUE"), col("F"))
    i_cs, i_cou = col("CLOSING SPREAD"), col("CLOSING O/U")
    i_os, i_oou = col("OPENING SPREAD"), col("OPENING O/U")
    if None in (i_gid, i_date, i_team, i_venue, i_f):
        print("  ! bigdataball: unrecognized column layout, skipping", file=sys.stderr)
        return pd.DataFrame()

    def to_iso(d):
        if isinstance(d, dt.datetime):
            return d.date().isoformat()
        if isinstance(d, dt.date):
            return d.isoformat()
        m, dd, y = str(d).split("/")
        return f"{y}-{int(m):02d}-{int(dd):02d}"

    def to_num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    games: dict[str, dict] = {}
    for r in rows[1:]:
        if not r or r[i_gid] is None:
            continue
        gid = str(r[i_gid])
        abbr = FULLNAME_TO_ABBR.get(r[i_team])
        if abbr is None:
            continue  # exhibition / national-team / unmapped row
        g = games.setdefault(gid, {"date": to_iso(r[i_date])})
        spread = to_num(r[i_cs]) if i_cs is not None else None
        if spread is None and i_os is not None:
            spread = to_num(r[i_os])
        total = to_num(r[i_cou]) if i_cou is not None else None
        if total is None and i_oou is not None:
            total = to_num(r[i_oou])
        if r[i_venue] == "Home":
            g["home"] = abbr
            g["home_score"] = to_num(r[i_f])
            g["spread_home"] = spread  # home row's own spread is already home-signed
        else:
            g["away"] = abbr
            g["away_score"] = to_num(r[i_f])
        if total is not None:
            g["total"] = total

    out = [dict(bdb_id=gid, **g) for gid, g in games.items()
           if "home" in g and "away" in g]
    return pd.DataFrame(out)


def ingest_bigdataball(games: pd.DataFrame, bdb: pd.DataFrame) -> pd.DataFrame:
    """Overlay BigDataBall's real spreads/totals (and backfill any missing
    scores) onto the existing games table, matched by date+home+away."""
    if bdb.empty:
        return games
    df = games.set_index("game_id")
    by_pair: dict[tuple, list] = {}
    for gid, row in df.iterrows():
        by_pair.setdefault((row.home, row.away), []).append((row.date, gid))

    def find_gid(date_str, home, away):
        target = dt.date.fromisoformat(date_str)
        best_gid, best_diff = None, None
        for cand_date, cand_gid in by_pair.get((home, away), []):
            diff = abs((dt.date.fromisoformat(cand_date) - target).days)
            if diff <= 1 and (best_diff is None or diff < best_diff):
                best_gid, best_diff = cand_gid, diff
        return best_gid

    for _, r in bdb.iterrows():
        gid = find_gid(r.date, r.home, r.away)
        if gid is not None:
            if pd.notna(r.get("spread_home")):
                df.loc[gid, "spread_home"] = r["spread_home"]
            if pd.notna(r.get("total")):
                df.loc[gid, "total"] = r["total"]
            if pd.isna(df.loc[gid, "home_score"]) and pd.notna(r.get("home_score")):
                df.loc[gid, "home_score"] = r["home_score"]
            if pd.isna(df.loc[gid, "away_score"]) and pd.notna(r.get("away_score")):
                df.loc[gid, "away_score"] = r["away_score"]
        else:
            new_gid = f"bdb_{r['bdb_id']}"
            df.loc[new_gid] = {
                "date": r["date"], "home": r["home"], "away": r["away"],
                "home_score": r.get("home_score"), "away_score": r.get("away_score"),
                "spread_home": r.get("spread_home"), "total": r.get("total"),
            }
    return (df.reset_index().rename(columns={"index": "game_id"})
              .sort_values("date").reset_index(drop=True))


def ingest_bigdataball_dir(games: pd.DataFrame) -> pd.DataFrame:
    if not os.path.isdir(BDB_DIR):
        return games
    found_any = False
    for fn in sorted(os.listdir(BDB_DIR)):
        if fn.lower().endswith((".xlsx", ".xlsm")):
            found_any = True
            fp = os.path.join(BDB_DIR, fn)
            print(f"Ingesting BigDataBall file: {fn}")
            try:
                bdb = parse_bigdataball_xlsx(fp)
                before = games.dropna(subset=["spread_home", "total"]).shape[0]
                games = ingest_bigdataball(games, bdb)
                after = games.dropna(subset=["spread_home", "total"]).shape[0]
                print(f"  parsed {len(bdb)} games; market lines "
                      f"{before} -> {after}")
                if after <= before:
                    print("  !! WARNING: no new lines were merged — check that "
                          "team names/dates in the file match.")
            except ImportError:
                print("  !! ERROR: openpyxl is not installed, cannot read .xlsx. "
                      "Add 'openpyxl' to the workflow's pip install step.",
                      file=sys.stderr)
            except Exception as e:  # noqa: BLE001
                print(f"  !! ERROR ingesting {fn}: {e}", file=sys.stderr)
    if not found_any:
        print("No BigDataBall .xlsx found in data/bigdataball/ "
              "(model will use game results only).")
    return games


# ----------------------------------------------------------------------------
# Model
# ----------------------------------------------------------------------------
def build_rankings(games: pd.DataFrame) -> tuple[pd.DataFrame, float, float]:
    """Weighted regression on expected/actual team points.

    Row per team-game:  pts = mu + o_team - d_opp + hca * (+.5 home / -.5 away)
    Market rows use line-implied points; result rows use actual points.
    Returns (rankings df, mu, hca).
    """
    g = games.copy()
    g["date"] = pd.to_datetime(g["date"])
    g = g.sort_values("date").reset_index(drop=True)

    completed = g.dropna(subset=["home_score", "away_score"])
    teams = sorted(set(g["home"]) | set(g["away"]))
    tidx = {t: i for i, t in enumerate(teams)}
    n = len(teams)

    # elapsed games per team, counted over completed games (most recent = 0)
    def elapsed_map(team: str) -> dict:
        tg = completed[(completed.home == team) | (completed.away == team)]
        ids = list(tg.sort_values("date")["game_id"])
        total = len(ids)
        return {gid: total - 1 - k for k, gid in enumerate(ids)}, total

    emaps = {t: elapsed_map(t) for t in teams}

    def elapsed(team: str, gid: str) -> int:
        emap, total = emaps[team]
        # upcoming game (not in completed set): nothing has elapsed since it
        return emap.get(gid, 0)

    rows_X, rows_y, rows_w = [], [], []

    def add_row(team, opp, pts, is_home, weight):
        x = np.zeros(2 + 2 * n)
        x[0] = 1.0                       # mu
        x[1] = 0.5 if is_home else -0.5  # hca
        x[2 + tidx[team]] = 1.0          # offense of team
        x[2 + n + tidx[opp]] = -1.0      # defense of opponent
        rows_X.append(x)
        rows_y.append(pts)
        rows_w.append(weight)

    # Market rows: decompose spread+total into expected points for each side.
    lines = g.dropna(subset=["spread_home", "total"])
    for _, r in lines.iterrows():
        margin = -r.spread_home              # expected home margin
        eh = (r.total + margin) / 2.0
        ea = (r.total - margin) / 2.0
        add_row(r.home, r.away, eh, True,
                1.0 / (elapsed(r.home, r.game_id) + ODDS_DENOM))
        add_row(r.away, r.home, ea, False,
                1.0 / (elapsed(r.away, r.game_id) + ODDS_DENOM))

    # Result rows: actual scores, heavily down-weighted.
    for _, r in completed.iterrows():
        add_row(r.home, r.away, r.home_score, True,
                1.0 / (elapsed(r.home, r.game_id) + RESULT_DENOM))
        add_row(r.away, r.home, r.away_score, False,
                1.0 / (elapsed(r.away, r.game_id) + RESULT_DENOM))

    X = np.array(rows_X)
    y = np.array(rows_y)
    w = np.array(rows_w)

    # Identifiability: sum(o) = 0 and sum(d) = 0, imposed as heavy pseudo-rows.
    big = 1e6
    c1 = np.zeros(2 + 2 * n); c1[2:2 + n] = 1.0
    c2 = np.zeros(2 + 2 * n); c2[2 + n:] = 1.0
    X = np.vstack([X, c1, c2])
    y = np.concatenate([y, [0.0, 0.0]])
    w = np.concatenate([w, [big, big]])

    sw = np.sqrt(w)
    beta, *_ = np.linalg.lstsq(X * sw[:, None], y * sw, rcond=None)
    mu, hca = beta[0], beta[1]
    o = beta[2:2 + n]
    d = beta[2 + n:]

    # Win-loss and SRS (simple rating: margin adjusted for opponent, iterated)
    wl, margins, opps = {}, {t: [] for t in teams}, {t: [] for t in teams}
    for t in teams:
        tg = completed[(completed.home == t) | (completed.away == t)]
        wins = ((tg.home == t) & (tg.home_score > tg.away_score)).sum() + \
               ((tg.away == t) & (tg.away_score > tg.home_score)).sum()
        wl[t] = (int(wins), int(len(tg) - wins))
        for _, r in tg.iterrows():
            if r.home == t:
                margins[t].append(r.home_score - r.away_score)
                opps[t].append(r.away)
            else:
                margins[t].append(r.away_score - r.home_score)
                opps[t].append(r.home)
    srs = {t: (np.mean(margins[t]) if margins[t] else 0.0) for t in teams}
    for _ in range(200):
        srs = {t: (np.mean(margins[t]) if margins[t] else 0.0)
               + (np.mean([srs[o_] for o_ in opps[t]]) if opps[t] else 0.0)
               for t in teams}
    srs_vals = np.array([srs[t] for t in teams])
    srs = {t: srs[t] - srs_vals.mean() for t in teams}

    out = pd.DataFrame({
        "team": teams,
        "gpf": o + d,
        "ogpf": o,
        "dgpf": d,
        "gou": 2 * mu + o - d,
        "srs": [srs[t] for t in teams],
        "w": [wl[t][0] for t in teams],
        "l": [wl[t][1] for t in teams],
    }).sort_values("gpf", ascending=False).reset_index(drop=True)
    out["rank"] = out.index + 1
    for col in ("ogpf", "dgpf", "gou", "srs"):
        out[f"{col}_rank"] = out[col].rank(ascending=False, method="min").astype(int)
    return out, mu, hca


# ----------------------------------------------------------------------------
# History + rendering
# ----------------------------------------------------------------------------
def append_history(rankings: pd.DataFrame, asof: str) -> pd.DataFrame:
    cols = ["date", "team", "rank", "gpf", "ogpf", "dgpf", "gou"]
    hist = pd.read_csv(HISTORY_CSV) if os.path.exists(HISTORY_CSV) \
        else pd.DataFrame(columns=cols)
    hist = hist[hist.date != asof]  # idempotent per day
    snap = rankings[["team", "rank", "gpf", "ogpf", "dgpf", "gou"]].copy()
    snap.insert(0, "date", asof)
    hist = pd.concat([hist, snap], ignore_index=True).sort_values(["date", "rank"])
    hist.to_csv(HISTORY_CSV, index=False)
    return hist


def sparkline(vals: list[float], color="#1f6feb", w=88, h=30) -> str:
    vals = vals[-SPARK_N:]
    if len(vals) < 2:
        return ""
    lo, hi = min(vals), max(vals)
    rng = (hi - lo) or 1.0
    pts = []
    for i, v in enumerate(vals):
        x = 3 + i * (w - 6) / (len(vals) - 1)
        yy = h - 4 - (v - lo) / rng * (h - 8)
        pts.append(f"{x:.1f},{yy:.1f}")
    last = pts[-1].split(",")
    return (f'<svg class="spark" viewBox="0 0 {w} {h}" width="{w}" height="{h}">'
            f'<polyline points="{" ".join(pts)}" fill="none" stroke="{color}" '
            f'stroke-width="1.6"/>'
            f'<circle cx="{last[0]}" cy="{last[1]}" r="2.2" fill="{color}"/></svg>')


def last_week_rank(hist: pd.DataFrame, team: str, asof: str):
    target = (dt.date.fromisoformat(asof) - dt.timedelta(days=7)).isoformat()
    old = hist[(hist.team == team) & (hist.date <= target)]
    if len(old):
        return int(old.sort_values("date").iloc[-1]["rank"])
    return None


def render_html(rankings, hist, asof, mu, hca, n_lines, n_results) -> str:
    def series(team, col):
        s = hist[hist.team == team].sort_values("date")[col].tolist()
        return s

    rows = []
    for _, r in rankings.iterrows():
        lw = last_week_rank(hist, r.team, asof)
        lw_txt = "&ndash;" if lw is None else str(lw)
        move = ""
        if lw is not None and lw != r["rank"]:
            up = lw > r["rank"]
            move = (f'<span class="mv {"up" if up else "dn"}">'
                    f'{"▲" if up else "▼"}{abs(lw - r["rank"])}</span>')
        rows.append(f"""
      <tr>
        <td class="rk">{r['rank']}</td>
        <td class="tm"><span class="ab">{r.team}</span>
            <span class="nm">{TEAM_NAMES.get(r.team, r.team)}</span></td>
        <td class="lw">{lw_txt} {move}</td>
        <td class="num main">{r.gpf:+.1f}</td>
        <td class="sp">{sparkline(series(r.team, 'gpf'))}</td>
        <td class="num">{r.ogpf:+.1f} <span class="sub">({r.ogpf_rank})</span></td>
        <td class="num">{r.dgpf:+.1f} <span class="sub">({r.dgpf_rank})</span></td>
        <td class="num">{r.gou:.1f} <span class="sub">({r.gou_rank})</span></td>
        <td class="num">{r.srs:+.1f} <span class="sub">({r.srs_rank})</span></td>
        <td class="num">{r.w}-{r.l}</td>
      </tr>""")

    return f"""<!doctype html>
<html lang="en"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WNBA Betting Market Rankings</title>
<style>
  :root {{
    --ink:#14212e; --sub:#6b7a89; --line:#dfe6ec; --acc:#1f6feb;
    --bg:#f7f9fb; --card:#ffffff;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--bg); color:var(--ink);
         font:15px/1.45 "Helvetica Neue", Arial, sans-serif; }}
  .wrap {{ max-width:980px; margin:0 auto; padding:28px 18px 60px; }}
  h1 {{ font-size:26px; margin:0 0 2px; letter-spacing:-.01em; }}
  .asof {{ color:var(--sub); margin:0 0 18px; }}
  table {{ width:100%; border-collapse:collapse; background:var(--card);
           border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
  th {{ text-align:right; font-size:12px; text-transform:uppercase;
        letter-spacing:.06em; color:var(--sub); padding:10px 10px;
        border-bottom:2px solid var(--line); background:#fcfdfe; }}
  th.l {{ text-align:left; }}
  td {{ padding:8px 10px; border-bottom:1px solid var(--line);
        text-align:right; white-space:nowrap; }}
  tr:last-child td {{ border-bottom:none; }}
  tr:hover td {{ background:#f2f7ff; }}
  .rk {{ color:var(--sub); width:34px; }}
  .tm {{ text-align:left; }}
  .ab {{ font-weight:700; }}
  .nm {{ color:var(--sub); font-size:13px; margin-left:6px; }}
  .lw {{ color:var(--sub); }}
  .mv.up {{ color:#1a7f37; font-size:12px; }}
  .mv.dn {{ color:#c1341f; font-size:12px; }}
  .num.main {{ font-weight:700; font-size:16px; }}
  .sub {{ color:var(--sub); font-size:11px; }}
  .sp {{ padding:2px 6px; }}
  .spark {{ display:block; }}
  .foot {{ color:var(--sub); font-size:13px; margin-top:16px; }}
  .foot b {{ color:var(--ink); }}
  @media (max-width:720px) {{ .nm, .sp {{ display:none; }} }}
</style></head>
<body><div class="wrap">
  <h1>WNBA Betting Market Rankings</h1>
  <p class="asof">As of {asof} &middot; implied by point spreads and totals
     &middot; updates daily 5:00 AM ET</p>
  <table>
    <thead><tr>
      <th>Rk</th><th class="l">Team</th><th>LstWk</th>
      <th>GPF</th><th>Trend</th><th>oGPF</th><th>dGPF</th>
      <th>GOU</th><th>SRS</th><th>W-L</th>
    </tr></thead>
    <tbody>{''.join(rows)}
    </tbody>
  </table>
  <p class="foot">
    <b>GPF</b>: generic points favored vs. an average opponent, neutral court.
    <b>oGPF/dGPF</b>: offensive / defensive components (dGPF &gt; 0 = good defense).
    <b>GOU</b>: implied over/under vs. an average opponent.
    Weighted regression on {n_lines} market lines and {n_results} game results;
    line weight 1/(elapsed+{ODDS_DENOM}), result weight 1/(elapsed+{RESULT_DENOM}).
    Estimated home court advantage: <b>{hca:.2f}</b> pts.
    League average score: <b>{mu:.1f}</b>.
    Methodology per
    <a href="https://www.inpredictable.com/2016/07/betting-market-rankings-for-wnba.html">
    inpredictable</a>.
  </p>
</div></body></html>"""


# ----------------------------------------------------------------------------
# Synthetic test
# ----------------------------------------------------------------------------
def run_test():
    rng = np.random.default_rng(7)
    teams = list(TEAM_NAMES)
    true_o = rng.normal(0, 3, len(teams))
    true_d = rng.normal(0, 3, len(teams))
    true_o -= true_o.mean(); true_d -= true_d.mean()
    mu_true, hca_true = 82.0, 2.8
    rows, gid = [], 0
    start = dt.date(2026, 5, 15)
    for day in range(60):
        order = rng.permutation(len(teams))
        for k in range(0, len(teams) - 1, 2):
            i, j = order[k], order[k + 1]
            eh = mu_true + true_o[i] - true_d[j] + hca_true / 2
            ea = mu_true + true_o[j] - true_d[i] - hca_true / 2
            rows.append({
                "game_id": str(gid), "date": (start + dt.timedelta(days=day)).isoformat(),
                "home": teams[i], "away": teams[j],
                "home_score": round(eh + rng.normal(0, 10)),
                "away_score": round(ea + rng.normal(0, 10)),
                "spread_home": round(-((eh - ea) + rng.normal(0, .5)) * 2) / 2,
                "total": round(((eh + ea) + rng.normal(0, 1)) * 2) / 2,
            }); gid += 1
    games = pd.DataFrame(rows)
    rk, mu, hca = build_rankings(games)
    true_gpf = {t: true_o[i] + true_d[i] for i, t in enumerate(teams)}
    est = rk.set_index("team")["gpf"]
    err = np.mean([abs(est[t] - true_gpf[t]) for t in teams])
    corr = np.corrcoef([est[t] for t in teams], [true_gpf[t] for t in teams])[0, 1]
    print(f"recovered mu={mu:.2f} (true {mu_true}), hca={hca:.2f} (true {hca_true})")
    print(f"GPF mean abs error={err:.3f}, correlation={corr:.4f}")
    assert corr > 0.95 and abs(hca - hca_true) < 0.75 and abs(mu - mu_true) < 1.0
    print("TEST PASSED")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--backfill", action="store_true",
                    help="re-scan every date from season start")
    ap.add_argument("--no-fetch", action="store_true")
    ap.add_argument("--test", action="store_true")
    args = ap.parse_args()

    if args.test:
        run_test(); return

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(OUT_DIR, exist_ok=True)

    games = load_games()
    today_et = dt.datetime.now(EASTERN).date()

    if not args.no_fetch:
        if args.backfill or games.empty:
            start = dt.date.fromisoformat(SEASON_START)
        else:
            last = pd.to_datetime(games["date"]).max().date()
            start = last - dt.timedelta(days=2)  # small overlap to catch late finals
        print(f"Fetching {start} .. {today_et}")
        games = merge_games(games, fetch_range(start, today_et))

    games = ingest_bigdataball_dir(games)
    games.to_csv(GAMES_CSV, index=False)

    usable = games.dropna(subset=["spread_home", "total"])
    finals = games.dropna(subset=["home_score", "away_score"])
    if len(usable) + len(finals) < 20:
        print("Not enough data to model yet.", file=sys.stderr)
        sys.exit(1)

    rankings, mu, hca = build_rankings(games)
    asof = today_et.isoformat()
    hist = append_history(rankings, asof)

    html = render_html(rankings, hist, asof, mu, hca, len(usable), len(finals))
    with open(os.path.join(OUT_DIR, "index.html"), "w") as f:
        f.write(html)
    rankings.to_csv(os.path.join(OUT_DIR, "rankings.csv"), index=False)
    print(rankings[["rank", "team", "gpf", "ogpf", "dgpf", "gou", "srs"]]
          .to_string(index=False))
    print(f"\nmu={mu:.2f}  hca={hca:.2f}  lines={len(usable)}  results={len(finals)}")


if __name__ == "__main__":
    main()
